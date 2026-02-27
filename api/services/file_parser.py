"""
file_parser.py - 文件解析工具模块

支持CSV、Excel(xlsx)、JSON格式文件的内存解析。
"""

import csv
import io
import json
from typing import Union

from loguru import logger


def parse_csv(content: Union[str, bytes]) -> list[dict]:
    """
    解析CSV文件内容。

    Args:
        content: CSV文件内容（字符串或字节）

    Returns:
        解析后的字典列表

    Raises:
        ValueError: CSV格式无效
    """
    if isinstance(content, bytes):
        # 尝试多种编码
        for encoding in ["utf-8", "gbk", "gb2312", "utf-8-sig"]:
            try:
                content = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("无法识别CSV文件编码，请使用UTF-8编码")

    if not content.strip():
        raise ValueError("CSV文件内容为空")

    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)

    if not rows:
        raise ValueError("CSV文件没有数据行")

    return rows


def parse_excel(content: bytes) -> list[dict]:
    """
    解析Excel(xlsx)文件内容。

    Args:
        content: Excel文件的字节内容

    Returns:
        解析后的字典列表

    Raises:
        ValueError: Excel格式无效
        ImportError: openpyxl未安装
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl库未安装，无法解析Excel文件")

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    except Exception as e:
        raise ValueError(f"Excel文件格式无效: {e}")

    ws = wb.active
    if ws is None:
        raise ValueError("Excel文件没有活动工作表")

    rows_iter = ws.iter_rows(values_only=True)

    # 第一行作为表头
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel文件为空")

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers)]

    rows = []
    for row in rows_iter:
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(value).strip() if value is not None else ""
        if any(v for v in row_dict.values()):  # 跳过全空行
            rows.append(row_dict)

    wb.close()

    if not rows:
        raise ValueError("Excel文件没有数据行")

    return rows


def parse_json(content: Union[str, bytes]) -> dict:
    """
    解析JSON文件内容。

    Args:
        content: JSON文件内容（字符串或字节）

    Returns:
        解析后的字典

    Raises:
        ValueError: JSON格式无效
    """
    if isinstance(content, bytes):
        try:
            content = content.decode("utf-8")
        except UnicodeDecodeError:
            content = content.decode("utf-8-sig")

    if not content.strip():
        raise ValueError("JSON文件内容为空")

    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式无效: {e}")

    return data


def detect_file_type(filename: str) -> str:
    """
    根据文件名检测文件类型。

    Args:
        filename: 文件名

    Returns:
        文件类型: "csv", "excel", "json", "unknown"
    """
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        return "csv"
    elif name_lower.endswith((".xlsx", ".xls")):
        return "excel"
    elif name_lower.endswith(".json"):
        return "json"
    else:
        return "unknown"


def validate_file(filename: str, content: bytes, allowed_types: list[str], max_size_mb: float = 10.0) -> str:
    """
    校验上传文件的类型和大小。

    Args:
        filename: 文件名
        content: 文件内容字节
        allowed_types: 允许的文件类型列表 ["csv", "excel", "json"]
        max_size_mb: 最大文件大小（MB）

    Returns:
        错误信息，为空字符串则表示校验通过

    """
    file_type = detect_file_type(filename)
    if file_type == "unknown":
        return f"不支持的文件格式: {filename}，请上传 {', '.join(allowed_types)} 格式文件"

    if file_type not in allowed_types:
        return f"该接口不支持 {file_type} 格式，请上传 {', '.join(allowed_types)} 格式文件"

    size_mb = len(content) / (1024 * 1024)
    if size_mb > max_size_mb:
        return f"文件过大 ({size_mb:.1f}MB)，最大允许 {max_size_mb}MB"

    return ""
